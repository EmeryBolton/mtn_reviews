import openpyxl
import csv
import warnings
from pathlib import Path

warnings.filterwarnings('ignore')

data_dir = Path(__file__).parent
wb = openpyxl.load_workbook(data_dir / 'ski_resorts_all.xlsx', read_only=True, data_only=True)

# --- 1. Geocoding sheet (addresses + lat/lng) ---
geo_rows = []
ws = wb['Geocoding']
for row in ws.iter_rows(values_only=True):
    if row[0] and row[0] != 'addresskey':
        geo_rows.append({
            'address': str(row[0]).strip(),
            'latitude': str(row[1]).strip() if row[1] else '',
            'longitude': str(row[2]).strip() if row[2] else '',
            'flag': str(row[3]).strip() if len(row) > 3 and row[3] else '',
        })

with open(data_dir / 'geocoded_addresses.csv', 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['address', 'latitude', 'longitude', 'flag'])
    w.writeheader()
    w.writerows(geo_rows)
print(f"Geocoding: {len(geo_rows)} addresses saved")

# --- 2. Canada inventory (all non-empty rows) ---
canada_rows = []
ws = wb['All Canada Ski Areas (Incomplet']
headers = None
for i, row in enumerate(ws.iter_rows(values_only=True)):
    vals = [str(v).strip() if v is not None else '' for v in row]
    if not any(vals):
        continue
    if headers is None and i < 20:
        # Try to find header row
        if any(v for v in vals[:3]):
            if i > 0:  # skip title row
                headers = vals
                continue
        continue
    if headers and any(vals):
        canada_rows.append(dict(zip(headers, vals)))

if canada_rows:
    with open(data_dir / 'canada_raw.csv', 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(canada_rows)
    print(f"Canada: {len(canada_rows)} rows saved")
else:
    print("Canada sheet: no data rows found (may be formula-driven or empty)")

# --- 3. Quebec sheet full scan ---
quebec_data = []
ws = wb['Quebec Ski Areas Inventory']
all_rows = list(ws.iter_rows(values_only=True))
print(f"Quebec sheet actual row count: {len(all_rows)}")
# Print first 20 rows to a file for inspection
with open(data_dir / 'quebec_inspect.txt', 'w', encoding='utf-8') as f:
    for i, row in enumerate(all_rows[:30]):
        f.write(f"row {i+1}: {row}\n")

# --- 4. MASTERSHEET scan ---
ws = wb['MASTERSHEET']
ms_rows = list(ws.iter_rows(values_only=True))
print(f"MASTERSHEET row count: {len(ms_rows)}")
with open(data_dir / 'mastersheet_inspect.txt', 'w', encoding='utf-8') as f:
    for i, row in enumerate(ms_rows[:20]):
        f.write(f"row {i+1}: {row}\n")

print("Done.")
